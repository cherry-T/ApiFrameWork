from openpyxl import load_workbook

from common import project_path
from common.common_fun import CommunFun
from common.do_info import Do_Info
from common.do_mysql import DoMysql
from common.my_log import MyLog
from common.read_config import ConfigRead
from common.common_fun import CommunFun
from Common.data_manage import DataManage
import pandas as pd
import string

# 处理测试数据的模块(从excel获取，批量写回测试结果)
mylog = MyLog()
commonFun = CommunFun()


class DoExcel:

    def __init__(self, file_name, apiconfig):
        '''
        初始化函数，用于默认选择了接口用例文件和读取用例执行情况的配置文件
        filename:需要传入excel或其他文件的地址
        apiconfig: 获取配置文件，配置哪个sheet的内容,这个apiconfig一定在配置文件有的section
        '''
        self.file_name = file_name
        # 读取ApiConfig配置文件，里面配置各个模块要执行的用例
        self.sheet_list = eval(ConfigRead().read_config(project_path.config_path, apiconfig, 'sheetlist'))
        # print(self.sheet_list)

    def get_data(self):
        """
        读取excel接口测试用例的测试数据，如：id、模块、用例描述、请求方式、请求地址、请求用户数据、预期结果等
        根据读取api.config配置文件，来判断是否全部执行和指定用例执行
        :return:存放测试数据的列表，列表里面存放的是字典
        """
        # 打开excel文件
        wb = load_workbook(self.file_name)

        # 存放取出来的数据, 以列表存放字典[{"id":"1",..},{"id":"2",..}]
        test_data = []

        # 从配置文件读取要读取的sheet和具体的列，apiconfig
        # 控制读那个sheet表，循环遍历sheet，逐个sheet取数据，可以理解为不同的模块 apiconfig文件传入有两个模块的数据，这里也会一起获取的到
        for key in self.sheet_list:
            sheet = wb[key]
            # 获取最大行数
            # max_r = sheet.max_row
            # 读取配置api.config配置文件，配置all，即执行这个模块的所有用例
            if self.sheet_list[key] == 'all':
                for i in range(2, sheet.max_row+1):   # sheet.max_row sheet.max_column
                    sub_data = {}
                    sub_data['id'] = sheet.cell(i, 1).value
                    sub_data['module'] = sheet.cell(i, 2).value
                    sub_data['function'] = sheet.cell(i, 3).value
                    sub_data['title'] = sheet.cell(i, 4).value
                    sub_data['http_method'] = sheet.cell(i, 5).value
                    sub_data['url'] = sheet.cell(i, 6).value
                    sub_data['parm'] = sheet.cell(i, 7).value
                    sub_data['assertType'] = sheet.cell(i, 8).value
                    sub_data['ExpectedResult'] = sheet.cell(i, 9).value
                    # url：根据配置文件，动态替换host地址
                    # mylog.info("开始替换第{0}条_用例_{1}_的参数".format(sub_data['id'], sub_data['title']))
                    res_url = self.replace_url(sub_data['url'])
                    sub_data['url'] = res_url

                    # 要放在和判断是否有替换的同级，然后将所有模块的用例的添加的字典里面
                    test_data.append(sub_data)
            else:
                sheet_list = self.judge_getdata_by_testsuite(sheet_key=key)
                # 读取配置api.config配置文件，根据配置的id来执行
                for i in sheet_list:
                    sub_data = {}
                    # 循环遍历的行数 与 excel实际行数是+1的关系，所以要加1
                    sub_data['id'] = sheet.cell(i+1, 1).value
                    sub_data['module'] = sheet.cell(i+1, 2).value
                    sub_data['function'] = sheet.cell(i+1, 3).value
                    sub_data['title'] = sheet.cell(i+1, 4).value
                    sub_data['http_method'] = sheet.cell(i+1, 5).value
                    sub_data['url'] = sheet.cell(i+1, 6).value
                    sub_data['parm'] = sheet.cell(i + 1, 7).value
                    sub_data['assertType'] = sheet.cell(i + 1, 8).value
                    sub_data['ExpectedResult'] = sheet.cell(i+1, 9).value

                    # 将过去的sheet_name 保存下来
                    sub_data['sheet_name'] = key
                    # url：根据配置文件，动态替换host地址
                    res_url = self.replace_url(sub_data['url'])
                    sub_data['url'] = res_url

                    test_data.append(sub_data)
        mylog.info("读取excel数据完成")
        # 要在for循环完成之后，才能返回；否则，tab进一个，会导致只循环了一个sheet就退出循环，没有继续读取下一个sheet的all
        return test_data

    def get_data_name(self):
        """
        用于足力健批量签约，获取第一列的数据
        :return:存放测试数据的列表，列表里面存放的是字典
        """
        # 打开excel文件
        wb = load_workbook(self.file_name)

        # 存放取出来的数据, 以列表存放字典[{"id":"1",..},{"id":"2",..}]
        test_data = []

        # 从配置文件读取要读取的sheet和具体的列，apiconfig
        # 控制读那个sheet表，循环遍历sheet，逐个sheet取数据，可以理解为不同的模块
        for key in self.sheet_list:
            sheet = wb[key]
            # 获取最大行数
            # max_r = sheet.max_row
            # 读取配置api.config配置文件，配置all，即执行这个模块的所有用例
            if self.sheet_list[key] == 'all':
                for i in range(2, sheet.max_row+1):   # sheet.max_row sheet.max_column
                    sub_data = sheet.cell(i, 1).value
                    # 要放在和判断是否有替换的同级，然后将所有模块的用例的添加的字典里面
                    test_data.append(sub_data)
        mylog.info("读取excel数据完成")
        # 要在for循环完成之后，才能返回；否则，tab进一个，会导致只循环了一个sheet就退出循环，没有继续读取下一个sheet的all
        return test_data

    def get_data_original(self):
        """
        读取excel接口测试用例的测试数据，如：id、模块、用例描述、请求方式、请求地址、请求用户数据、预期结果等
        不需要做param的参数替换
        :return:存放测试数据的列表，列表里面存放的是字典
        """
        # 打开excel文件
        wb = load_workbook(self.file_name)

        # 存放取出来的数据
        test_data = []

        # 从配置文件读取要读取的sheet和具体的列，apiconfig
        # 控制读那个sheet表，循环遍历sheet，逐个sheet取数据，可以理解为不同的模块
        for key in self.sheet_list:
            sheet = wb[key]
            # 获取最大行数
            # max_r = sheet.max_row
            # 读取配置api.config配置文件，配置all，即执行这个模块的所有用例
            if self.sheet_list[key] == 'all':
                for i in range(2, sheet.max_row+1):   # sheet.max_row sheet.max_column
                    sub_data = {}
                    sub_data['id'] = sheet.cell(i, 1).value
                    sub_data['module'] = sheet.cell(i, 2).value
                    sub_data['function'] = sheet.cell(i, 3).value
                    sub_data['title'] = sheet.cell(i, 4).value
                    sub_data['http_method'] = sheet.cell(i, 5).value
                    sub_data['url'] = sheet.cell(i, 6).value
                    sub_data['activity_type'] = sheet.cell(i, 7).value
                    sub_data['parm'] = sheet.cell(i, 8).value
                    sub_data['assertType'] = sheet.cell(i, 9).value
                    sub_data['ExpectedResult'] = sheet.cell(i, 10).value
                    # ExpectedResult

                    if sub_data['id'] == 1:
                        # url：根据配置文件，动态替换host地址,配置为全部执行时，host都是一样，只需替换一次
                        mylog.info("开始替换第{0}条_用例_{1}_的参数".format(sub_data['id'], sub_data['title']))
                        res_url = self.replace_url(sub_data['url'])
                    sub_data['url'] = res_url

                    # 要放在和判断是否有替换的同级，然后将所有模块的用例的添加的字典里面
                    test_data.append(sub_data)

            elif self.sheet_list[key] == "half":
                # 获取当前excel文件总用例数的一半
                for i in range(2, int((sheet.max_row+1)/2)):
                    sub_data = {}
                    sub_data['id'] = sheet.cell(i, 1).value
                    sub_data['module'] = sheet.cell(i, 2).value
                    sub_data['function'] = sheet.cell(i, 3).value
                    sub_data['title'] = sheet.cell(i, 4).value
                    sub_data['http_method'] = sheet.cell(i, 5).value
                    sub_data['url'] = sheet.cell(i, 6).value
                    sub_data['activity_type'] = sheet.cell(i, 7).value
                    sub_data['parm'] = sheet.cell(i, 8).value
                    sub_data['assertType'] = sheet.cell(i, 9).value
                    sub_data['ExpectedResult'] = sheet.cell(i, 10).value

                    if sub_data['id'] == 1:
                        # url：根据配置文件，动态替换host地址,配置为全部执行时，host都是一样，只需替换一次
                        mylog.info("开始替换第{0}条_用例_{1}_的参数".format(sub_data['id'], sub_data['title']))
                        res_url = self.replace_url(sub_data['url'])
                    sub_data['url'] = res_url

                    # 要放在和判断是否有替换的同级，然后将所有模块的用例的添加的字典里面
                    test_data.append(sub_data)
            elif self.sheet_list[key] == "fifthpart":
                # 获取当前用例总数的1/5
                for i in range(2, int((sheet.max_row+1)/5)):
                    sub_data = {}
                    sub_data['id'] = sheet.cell(i, 1).value
                    sub_data['module'] = sheet.cell(i, 2).value
                    sub_data['function'] = sheet.cell(i, 3).value
                    sub_data['title'] = sheet.cell(i, 4).value
                    sub_data['http_method'] = sheet.cell(i, 5).value
                    sub_data['url'] = sheet.cell(i, 6).value
                    sub_data['activity_type'] = sheet.cell(i, 7).value
                    sub_data['parm'] = sheet.cell(i, 8).value
                    sub_data['assertType'] = sheet.cell(i, 9).value
                    sub_data['ExpectedResult'] = sheet.cell(i, 10).value
                    if sub_data['id'] == 1:
                        # url：根据配置文件，动态替换host地址,配置为全部执行时，host都是一样，只需替换一次
                        mylog.info("开始替换第{0}条_用例_{1}_的参数".format(sub_data['id'], sub_data['title']))
                        res_url = self.replace_url(sub_data['url'])
                    sub_data['url'] = res_url

                    # 要放在和判断是否有替换的同级，然后将所有模块的用例的添加的字典里面
                    test_data.append(sub_data)
            else:
                # 读取配置api.config配置文件，配置文件配置的用例id
                # 1、如果是同一个接口，用同一个url,里面的url只需要替换一次
                for t in range(1):
                    url_info = sheet.cell(2, 6).value
                    res_url = self.replace_url(url_info)

                # 2、获取当前文件测试用例集[test_suite]表中是否有数据，有则，取配置好的用例集的id，否则，取配置的好的用例id，两者均不是，默认取第一条
                sheet_list = self.judge_getdata_by_testsuite(sheet_key=key)

                # 3、获取到sheet_list，开始循环遍历获取数据
                for i in sheet_list:
                    sub_data = {}
                    # 循环遍历的行数 与 excel实际行数是+1的关系，所以要加1
                    sub_data['id'] = sheet.cell(i+1, 1).value
                    sub_data['module'] = sheet.cell(i+1, 2).value
                    sub_data['function'] = sheet.cell(i+1, 3).value
                    sub_data['title'] = sheet.cell(i+1, 4).value
                    sub_data['http_method'] = sheet.cell(i+1, 5).value
                    sub_data['url'] = sheet.cell(i+1, 6).value
                    sub_data['activity_type'] = sheet.cell(i+1, 7).value
                    sub_data['parm'] = sheet.cell(i + 1, 8).value
                    sub_data['assertType'] = sheet.cell(i + 1, 9).value
                    sub_data['ExpectedResult'] = sheet.cell(i+1, 10).value

                    # 将过去的sheet_name 保存下来
                    sub_data['sheet_name'] = key
                    sub_data['url'] = res_url

                    test_data.append(sub_data)
        mylog.info("读取excel数据完成")
        # 要在for循环完成之后，才能返回；否则，tab进一个，会导致只循环了一个sheet就退出循环，没有继续读取下一个sheet的all
        return test_data

    def judge_getdata_by_testsuite(self, sheet_key):
        '''
        【判断是否取测试用例集的用例】根据配置文件里面配置的key_value值判断
        1、如果是测试用例集，如"smoke", 则从smoke配置的list中取用例（结合当前读取的excel里面的test_suite）
        2、如果配置[1,5],具体的list,直接从列表取
        3、两者都不是，默认取第一条
        :param sheet_key: 配置文件的key值名
        :return: list
        '''
        row_list = self.get_appoint_sheet_data(sheetname="test_suite")
        if row_list:
            for list_i in row_list:
                # 判断用例集的名称和配置文件的名称是否一致，一致，则取用例集配置的用例id
                if list_i["suite"] == self.sheet_list[sheet_key]:
                    sheet_list = eval(list_i["parm"])
                    break
                else:
                    sheet_list = self.sheet_list[sheet_key]
        else:
            if isinstance(self.sheet_list[sheet_key], list):
                sheet_list = self.sheet_list[sheet_key]
            else:
                mylog.exception("用例集sheet表没有数据，默认读取第一条用例")
                sheet_list = [1]
        return sheet_list

    def get_data_oms(self):
        """
        与get_data方法的区别是，get_data_oms：增加了platform scene1场景的等参数
        获取excel表格的数据
        :return:存放测试数据的列表，列表里面存放的是字典
        """
        # 打开excel文件
        wb = load_workbook(self.file_name)

        # 存放取出来的数据, 以列表存放字典[{"id":"1",..},{"id":"2",..}]
        test_data = []

        # 从配置文件读取要读取的sheet和具体的列，apiconfig
        # 控制读那个sheet表，循环遍历sheet，逐个sheet取数据，可以理解为不同的模块
        for key in self.sheet_list:
            sheet = wb[key]
            # 获取最大行数
            # max_r = sheet.max_row
            # 读取配置api.config配置文件，配置all，即执行这个模块的所有用例
            if self.sheet_list[key] == 'all':
                for i in range(2, sheet.max_row+1):   # sheet.max_row sheet.max_column
                    sub_data = {}
                    sub_data['id'] = sheet.cell(i, 1).value
                    sub_data['module'] = sheet.cell(i, 2).value
                    sub_data['function'] = sheet.cell(i, 3).value
                    sub_data['title'] = sheet.cell(i, 4).value
                    sub_data['http_method'] = sheet.cell(i, 5).value
                    sub_data['url'] = sheet.cell(i, 6).value
                    sub_data['platformId'] = sheet.cell(i, 7).value
                    sub_data['parm'] = sheet.cell(i, 8).value
                    sub_data['scene1'] = sheet.cell(i, 9).value
                    sub_data['scene2'] = sheet.cell(i, 10).value
                    sub_data['scene3'] = sheet.cell(i, 11).value
                    sub_data['unexpected_scene1'] = sheet.cell(i, 12).value
                    sub_data['unexpected_scene2'] = sheet.cell(i, 13).value
                    sub_data['assertType'] = sheet.cell(i, 14).value
                    sub_data['ExpectedResult'] = sheet.cell(i, 15).value
                    # url：根据配置文件，动态替换host地址
                    # mylog.info("开始替换第{0}条_用例_{1}_的参数".format(sub_data['id'], sub_data['title']))
                    res_url = self.replace_url(sub_data['url'])
                    sub_data['url'] = res_url

                    # 要放在和判断是否有替换的同级，然后将所有模块的用例的添加的字典里面
                    test_data.append(sub_data)
            else:
                # 读取配置api.config配置文件，根据配置的id来执行
                # 从excel文件里面判断当前配置文件的冒烟集名称和excel集合名称是否一样
                sheet_list = self.judge_getdata_by_testsuite(sheet_key=key)
                for i in sheet_list:
                    sub_data = {}
                    # 循环遍历的行数 与 excel实际行数是+1的关系，所以要加1
                    sub_data['id'] = sheet.cell(i+1, 1).value
                    sub_data['module'] = sheet.cell(i+1, 2).value
                    sub_data['function'] = sheet.cell(i+1, 3).value
                    sub_data['title'] = sheet.cell(i+1, 4).value
                    sub_data['http_method'] = sheet.cell(i+1, 5).value
                    sub_data['url'] = sheet.cell(i+1, 6).value
                    sub_data['platformId'] = sheet.cell(i+1, 7).value
                    sub_data['parm'] = sheet.cell(i+1, 8).value
                    sub_data['scene1'] = sheet.cell(i+1, 9).value
                    sub_data['scene2'] = sheet.cell(i+1, 10).value
                    sub_data['scene3'] = sheet.cell(i+1, 11).value
                    sub_data['unexpected_scene1'] = sheet.cell(i+1, 12).value
                    sub_data['unexpected_scene2'] = sheet.cell(i+1, 13).value
                    sub_data['assertType'] = sheet.cell(i+1, 14).value
                    sub_data['ExpectedResult'] = sheet.cell(i+1, 15).value

                    # 将过去的sheet_name 保存下来
                    sub_data['sheet_name'] = key
                    # url：根据配置文件，动态替换host地址
                    res_url = self.replace_url(sub_data['url'])
                    sub_data['url'] = res_url

                    test_data.append(sub_data)
        mylog.info("读取excel数据完成")
        # 要在for循环完成之后，才能返回；否则，tab进一个，会导致只循环了一个sheet就退出循环，没有继续读取下一个sheet的all
        return test_data

    def get_data_by_rowid(self, sheetname, rowid):
        """
        param sheetname: 表名，执行对应的表名
        param rowid: 行id，指定行id读取数据
        :return:存放测试数据的列表，列表里面存放的是字典
        """
        # 传入文件的路径，打开文件
        wb = load_workbook(self.file_name)

        # 存放取出来的数据, 以列表存放字典[{"id":"1",..},{"id":"2",..}]
        test_data = []

        # 控制读那个sheet表，循环遍历sheet，逐个sheet取数据，可以理解为不同的模块
        sheet = wb[sheetname]

        # 根据传入的rowid，遍历取数据
        for i in rowid:
            sub_data = {}
            # 循环遍历的行数 与 excel实际行数是+1的关系，所以要加1
            sub_data['id'] = sheet.cell(i+1, 1).value
            sub_data['module'] = sheet.cell(i+1, 2).value
            sub_data['function'] = sheet.cell(i+1, 3).value
            sub_data['title'] = sheet.cell(i+1, 4).value
            sub_data['http_method'] = sheet.cell(i+1, 5).value
            sub_data['url'] = sheet.cell(i+1, 6).value
            sub_data['platformId'] = sheet.cell(i+1, 7).value
            sub_data['parm'] = sheet.cell(i+1, 8).value
            sub_data['scene1'] = sheet.cell(i+1, 9).value
            sub_data['scene2'] = sheet.cell(i+1, 10).value
            sub_data['scene3'] = sheet.cell(i+1, 11).value
            sub_data['unexpected_scene1'] = sheet.cell(i+1, 12).value
            sub_data['unexpected_scene2'] = sheet.cell(i+1, 13).value
            sub_data['assertType'] = sheet.cell(i+1, 14).value
            sub_data['ExpectedResult'] = sheet.cell(i+1, 15).value

            # 将过去的sheet_name 保存下来
            sub_data['sheet_name'] = sheet
            # url：根据配置文件，动态替换host地址
            res_url = self.replace_url(sub_data['url'])
            sub_data['url'] = res_url

            test_data.append(sub_data)
        mylog.info("读取excel数据完成")
        # 要在for循环完成之后，才能返回；否则，tab进一个，会导致只循环了一个sheet就退出循环，没有继续读取下一个sheet的all
        return test_data

    def get_data_by_pandas(self):
        '''
        【读取数据】比excel好的地方，excel加了列，方法要同步改，pandas不用
        :param filename: 文件路径
        :param sheetname: 执行的sheetname
        :param apiconfig: 配置读取的数据，有读取全部和读取配置的（可配置在配置文件或直接传入列表）
        :return: 返回list嵌套字典
        '''
        res = []
        for key in self.sheet_list:   # 根据配置文件apiconfig获取
            # 读取excel数据时，要设置keep_default_na为False,否则，空值显示为nan
            source_data = pd.read_excel(self.file_name, sheet_name=key, keep_default_na=False)
            if self.sheet_list[key] == "all":
                for i in source_data.index:
                    res.append(source_data.iloc[i].to_dict())
            else:
                for i in self.sheet_list[key]:
                    # 直接读取配置的列表行数（从0开始第一行数据（除去列名），如果要和配置的caseid一致，则需要-1进行对应）
                    res.append(source_data.iloc[i-1].to_dict())

        return res

    def get_appoint_sheet_data(self, sheetname):
        '''
        【读取固定sheet表的数据】
        :param sheetname: 执行的sheetname
        :return: 返回list嵌套字典
        '''
        res = []

        # 读取excel数据时，要设置keep_default_na为False,否则，空值显示为nan
        source_data = pd.read_excel(self.file_name, sheet_name=sheetname, keep_default_na=False)
        for i in source_data.index:
            res.append(source_data.iloc[i].to_dict())

        return res

    def get_data_by_rowid_pandas(self, filename, sheet_name, rowid:list):
        '''
        【读取数据】 指定文件路径和sheetname, 根据传入的rowid获取数据
        :param filename: 文件路径
        :param sheetname: 执行的sheetname
        :param apiconfig: 配置读取的数据，有读取全部和读取配置的（可配置在配置文件或直接传入列表）
        :return: 返回list嵌套字典
        '''
        res = []
        # 读取excel数据时，要设置keep_default_na为False,否则，空值显示为nan
        source_data = pd.read_excel(filename, sheet_name=sheet_name, keep_default_na=False)

        for i in rowid:
            # 直接读取配置的列表行数（从0开始第一行数据（除去列名），如果要和配置的caseid一致，则需要-1进行对应）
            res.append(source_data.iloc[i-1].to_dict())

        return res

    def write_back(self, sheet_name, row_id, ActualResult, TestResult):
        """
        批量写回比对的结果回excel文件，包含实际结果，和预期和实际比对的结果
        :param sheet_name:sheet_name名，定位到具体的sheet
        :param row_id:第几行
        :param ActualResult:实际结果
        :param TestResult:测试结果：通过或结束
        """
        # 回写进来，传入module，名字要一样
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]

        # 第几列已经固定好，后面若调整了excel，这里需要调整
        # 计算促销，增加了一列活动类型，回写excel时要判断，回写到第几列
        if sheet_name == "promotion_demo" or sheet_name == "promotion" or sheet_name == "create_wholeact" \
                or sheet_name == "promotion_test":
            sheet.cell(row_id, 11).value = ActualResult
            sheet.cell(row_id, 12).value = TestResult
        elif sheet_name == 'oms' or sheet_name == 'oms_basic' or sheet_name == 'audit' or sheet_name == 'inv_basic':
            sheet.cell(row_id, 16).value = ActualResult
            sheet.cell(row_id, 17).value = TestResult
        elif sheet_name == 'inv':
            sheet.cell(row_id, 17).value = ActualResult
            sheet.cell(row_id, 18).value = TestResult
        else:
            sheet.cell(row_id, 10).value = ActualResult
            sheet.cell(row_id, 11).value = TestResult

        wb.save(self.file_name)

    def write_data_to_excel_by_pandas(self, sheetname, rowid: int, actual_res, test_res):
        '''
        【回写数据到excel】通过pandas回写指定的测试数据到excel表
        :param filename: 文件名，文件的路径
        :param sheetname: excel的sheetname
        :param rowid:  列
        :param actual_res: 实际结果
        :param test_res:  比对结果（前面比对完传入）
        :return:
        '''
        source_data = pd.read_excel(self.file_name, sheet_name=sheetname)  # 返回一个DataFrame对象，多维数据结构

        #  todo: 需要确定rowid
        # 单一赋值操作进行更新，除去列名，行数从0开始
        # 使用source_data["actual"][0]=actual_res，会返回SettingwithCopyWarning的警告
        source_data.loc[rowid, "actual"] = actual_res
        source_data.loc[rowid, "result"] = test_res

        # 数据写到excel里面，但是报错
        with pd.ExcelWriter(self.file_name) as writer:
            source_data.to_excel(writer, sheet_name=sheetname, index=False)


    def replace_parm(self, replaceparm, title=None, module=None):
        '''
        动态替换，excel的parm字段的内容，如果有包含指定的${replace}，则进行相应的替换
        :param
        replaceparm: 传进来要判断是否需要动态替换的参数
        title: 增加标题判断,设计到多个用例共用的时候，可增加title，再做分支判断
        :return:
        '''
        res_data = ""
        # 传入的替换的字符串内容，就是从读取excel中直接读取回来的value值
        # replace_parm = sheet.cell(i, 6).value
        memberid = CommunFun().create_memberId()
        if replaceparm.find('${reg_memberid}') != -1 and replaceparm.find('${mobile}') != -1:
            # update_one = replaceparm.replace('${reg_memberid}', str(memberid))
            # 随机生成的手机号
            mobile = CommunFun().create_mobile()
            param_t = commonFun.dynamic_create_dict(replaceparm)
            if "reg_memberid" in param_t.keys() or "mobile" in param_t.keys():
                param_t["reg_memberid"] = str(memberid)
                param_t["mobile"] = str(mobile)
            res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
        elif replaceparm.find('${oldMoblie}') != -1:
            # 修改电话接口
            res_data = self.replace_crm_mobile(replaceparm)
        elif replaceparm.find('${pin_activity_del}') != -1:
            # 查找已经删除的拼团活动id
            act_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                    ouid=str(Do_Info.ou_id), type=3)
            if act_info != 0:
                param_t = commonFun.dynamic_create_dict(replaceparm)
                if "pin_activity_del" in param_t.keys() or "join_team_id" in param_t.keys():
                    param_t["pin_activity_del"] = str(act_info[0])
                    param_t["join_team_id"] = str(act_info[1])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("查找不到已经删除的拼团活动，不进行参数替换！")
                res_data = replaceparm
        elif replaceparm.find('${pin_activity_stop}') != -1:
            # 查找已经结束的活动id
            pin_act = DoMysql('CRM').get_pinactivityid_and_teamid(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id), type=2)
            if pin_act != 0:
                res_data = replaceparm.replace('${pin_activity_stop}', str(pin_act[0]))
            else:
                mylog.error("查找不到已结束的拼团活动，不进行参数替换")
        elif replaceparm.find('${pin_activityid}') != -1:
            # 获取没有被删掉的拼团活动的id，用于查看拼团活动、拼团活动规则
            result = DoMysql('CRM').get_pinactivityid_and_teamid(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),type=0)
            if result != 0:
                param_t = commonFun.dynamic_create_dict(replaceparm)
                if "pin_activityid" in param_t.keys() or "pin_teamid" in param_t.keys() or "unionid" in param_t.keys():
                    param_t["pin_activityid"] = str(result[0])
                    param_t["pin_teamid"] = str(result[1])
                    if replaceparm.find('${unionid}') != -1:
                        userunionid = CommunFun().create_unionid()
                        param_t["unionid"] = userunionid
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("查找已审核状态的拼团活动失败，不进行参数替换！")
        elif replaceparm.find('${pin_list_union}') != -1:
            # 获取拼券列表接口,抽取逻辑到替换方法
            res_data = self.replace_pin_act_list(replaceparm)
        elif replaceparm.find('${pin_user_list_act}') != -1:
            # 获取拼团人员列表，抽取逻辑到替换方法
            res_data = self.replace_pin_act_user_list(replaceparm)
        elif replaceparm.find('${validate_coupon_act}') != -1:
            # 拼团活动领券验证
            res_data = self.replace_valivate_coupon_info(replaceparm)
        elif replaceparm.find('${create_pin_act}') != -1:
            # 拼团活动领券验证，拼团活动开团
            res_data = self.open_pin_activity_info(replaceparm)
        elif replaceparm.find('${join_team_id}') != -1:
            act_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                    ouid=str(Do_Info.ou_id), type=5)
            if act_info != 0:
                res_data = replaceparm.replace('${join_team_id}', str(act_info[1]))
            else:
                mylog.error("查找不到状态正常的拼团活动，不进行参数替换！")
                res_data = replaceparm
        elif replaceparm.find('${use_couponId}') != -1:
            # 门店派券活动、门店派券（获取优惠券详情）、卡券的id
            res_data = self.replace_coupon_shopissue(replaceparm)
        elif replaceparm.find('${use_coupon_code}') != -1:
            #  根据卡券码获取使用状态 卡券的编码 ccoupon_code
            coupon_code = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id), type=0)
            if coupon_code != 0:
                # 替换的卡券的编码
                res_data = replaceparm.replace('${use_coupon_code}', str(coupon_code[4]))
            else:
                mylog.error("查找不到正常使用的卡券编码，不进行参数替换！")
                res_data = replaceparm
        elif replaceparm.find('${del_use_couponId}') != -1:
            # 【卡券领券、卡券详情、根据卡券码获取使用状态】 已经删除的卡券的分支，couponid
            res_data = self.replace_del_coupon_id(replaceparm)
        elif replaceparm.find('${coupon_name}') != -1:
            # 用于获取非微信卡券的可用卡券 卡券名字，
            coupon_name = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id), type=5)
            res_data = replaceparm.replace('${coupon_name}', str(coupon_name[3]))
        elif replaceparm.find('${wx_coupon_name}') != -1:
            # 获取微信卡券的可用卡券，微信的卡券名
            coupon_name = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id), type=4)
            res_data = replaceparm.replace('${wx_coupon_name}', str(coupon_name[3]))
        elif replaceparm.find('${del_coupon_name}') != -1:
            # 获取微信卡券的可用卡券：已经删除的卡券，不区分微信和非微信
            del_coupon = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                      ouid=str(Do_Info.ou_id), type=6)
            if del_coupon != 0:
                res_data = replaceparm.replace('${del_coupon_name}', str(del_coupon[3]))
            else:
                mylog.error("查找不到已经删除的卡券")
                res_data = replaceparm
        elif replaceparm.find('${send_couponId}') != -1:
            # 用于卡券发送，会员领券接口 SendCouponToMemberForMall，替换卡券编码和会员信息
            # 获取正常的卡券
            res_data = self.replace_send_coupon_info(replaceparm)
        elif replaceparm.find('${invitation_activity}') != -1:
            # 邀请有礼邀请记录接口
            res_data = self.replace_inviation_activity(replaceparm)
        elif replaceparm.find('${baobao_crmid}') != -1:
            # 家庭档案-新增家庭档案，查找小于系统设置的最多添加的家庭参数的crm会员
            crm_id = DoMysql('CRM').get_SaveCrmMemberBaobao_memberid(groupid=str(Do_Info.group_id),
                                                                     ouid=str(Do_Info.ou_id), type=0)
            if crm_id != 0:
                res_data = replaceparm.replace('${baobao_crmid}', str(crm_id[0]))
            else:
                mylog.error("查不到crm会员家庭档案小于系统设置的人数失败，不进行参数替换")
                res_data = replaceparm
        elif replaceparm.find('${baobao_beyond_crmid}') != -1:
            # 查找已经超出系统设置的的人数的crm会员，不能再新增
            crm_id = DoMysql('CRM').get_SaveCrmMemberBaobao_memberid(groupid=str(Do_Info.group_id),
                                                                     ouid=str(Do_Info.ou_id), type=1)
            # 更新【最多可新增多少个家庭成员】参数最大的可新增人数为2
            param = {"groupid": str(Do_Info.group_id), "ouid": str(Do_Info.ou_id), "para_name": "最多可新增多少个家庭成员",
                     "para_value": '2', "show_value": '2'}
            DoMysql('BAS').update_bas_para_config(param=param)
            if crm_id != 0:
                res_data = replaceparm.replace('${baobao_beyond_crmid}', str(crm_id[0]))
            else:
                mylog.error("查不到crm会员家庭档案大于系统设置的人数失败，不进行参数替换")
                res_data = replaceparm
        elif replaceparm.find('${baobao_id}') != -1:
            # 用于更新家庭档案、查看档案详情
            baobao_info = DoMysql('CRM').get_crmbaobao(groupid=str(Do_Info.group_id),
                                                       ouid=str(Do_Info.ou_id), type=0)
            if baobao_info != 0:
                update_one = replaceparm.replace('${baobao_id}', str(baobao_info[0]))
                if update_one.find('${update_baobao_crmid}') != -1:
                    res_data = update_one.replace('${update_baobao_crmid}', str(baobao_info[1]))
                else:
                    res_data = update_one
            else:
                mylog.error("查不到crm会员家庭档案大于系统设置的人数失败，不进行参数替换")
                res_data = replaceparm
        elif replaceparm.find('${del_baobao_id}') != -1:
            # 用于更新家庭档案、查看档案详情
            baobao_info = DoMysql('CRM').get_crmbaobao(groupid=str(Do_Info.group_id),
                                                       ouid=str(Do_Info.ou_id), type=1)
            if baobao_info != 0:
                res_data = replaceparm.replace('${del_baobao_id}', str(baobao_info[0]))
            else:
                mylog.error("查不到已经删除的家庭档案大于系统设置的人数失败，不进行参数替换")
                res_data = replaceparm
        elif replaceparm.find('${introducermobile}') != -1 or \
                replaceparm.find('${not_weike_member_mobile}') != -1 or \
                replaceparm.find('${weike_member_mobile}') != -1 or \
                replaceparm.find('${handling_weike_mobile}') != -1 or \
                replaceparm.find('${refuse_weike_mobile}') != -1:
            # 微客分销：申请成为微客，检查手机号的审核状态、验证是否为会员的接口
            res_data = self.replace_weike_mobile(replaceparm)
        elif replaceparm.find('${weike_crmid}') != -1 or replaceparm.find('${expand_weike_crmid}') != -1 \
                or replaceparm.find('${no_expand_weike_crmid}') != -1 or replaceparm.find('${wait_check_crmid}') != -1 or \
                replaceparm.find('${not_weike_crmid}') != -1:
            # 微客分销：查看我的收益 收益明细 提现申请
            res_data = self.replace_weike_expand_info(replaceparm)
        elif replaceparm.find('${new_weike_crmid}') != -1 or replaceparm.find('${del_weike_crmid}') != -1:
            # 验证会员所属计佣微客： ${share_weike_id}分享人微客id，${del_weike_crmid}：已删除的会员id
            res_data = self.replace_check_weike_bymember(replaceparm)
        elif replaceparm.find('${extension_spu}') != -1:
            # 推广商品详情的spu
            spu_info = DoMysql('CRM').get_product_spu(groupid=str(Do_Info.group_id), shopid=str(Do_Info.shop_id))
            if spu_info != 0:
                res_data = replaceparm.replace('${extension_spu}', str(spu_info[0]))
            else:
                mylog.error("查找不到推广商品的spu,不进行参数替换")
                res_data = replaceparm
        elif replaceparm.find('${inspection_record}') != -1:
            # 验脚记录详情
            record_id = DoMysql('CRM').get_inspection_record(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                             shopid=str(Do_Info.shop_id))
            if record_id != 0:
                res_data = replaceparm.replace('${inspection_record}', str(record_id[0]))
            else:
                mylog.error("查不到{0}该店铺的验脚记录，不进行参数替换".format(Do_Info.shop_id))
                res_data = replaceparm
        elif replaceparm.find('${billno}') != -1 or replaceparm.find('${no_billno_card_code}') != -1:
            # # 提交消费单、提交消费记录、上传消费记录，bill_No、sbillNo单号,根据当前日期+3位随机数，随机生成单号
            res_data = self.replace_bill_info(replaceparm)
        elif replaceparm.find('${update_bill_status}') != -1:
            # 更新消费单状态：找出单据状态是1的消费单单号
            res_data = self.replace_update_bill_info(replaceparm)
        elif replaceparm.find('${mall_memberid}') != -1 or replaceparm.find('${crm_memberid}') != -1 \
                or replaceparm.find('${del_mall_memberid}') != -1 or replaceparm.find('${del_crm_member}') != -1:
            # ${mall_memberid}：商城的会员id，crm对接商城，很多接口需要用到这个字段
            # ${crm_memberid}: crm的会员id
            res_data = self.replace_mallmeber_crmid(replaceparm)
        elif replaceparm.find('${binding_card_mall_meb}') != -1:
            # 绑定会员卡的查出来的数据，有点问题，还没调试完成，会员卡号已0开头，转换的时候，会去掉0，导致找不到商城绑定信息
            res_data = self.replace_change_card_code(replaceparm)
        elif replaceparm.find('${releasecoupon_bymember}') != -1:
            # 【根据会员卡号或手机号释放所有占用的卡券】： ${releasecoupon_bymember}：通过会员卡号，
            release_meb = DoMysql('CRM').get_occupymorecoupon_membercard(type=0, groupid=str(Do_Info.group_id))
            if release_meb != 0:
                res_data = replaceparm.replace('${releasecoupon_bymember}', str(release_meb[0]))
            else:
                mylog.error("通过会员卡号，查找已占用的卡券，无数据，不进行参数替换")
        elif replaceparm.find('${nooccupycoupon_bymember}') != -1:
            # 【根据会员卡号或手机号释放所有占用的卡券】：没有占用卡券的会员卡号
            meb = DoMysql('CRM').get_occupymorecoupon_membercard(type=1, groupid=str(Do_Info.group_id))
            if meb != 0:
                res_data = replaceparm.replace('${nooccupycoupon_bymember}', str(meb[0]))
            else:
                mylog.error("查找不到没有占用的会员卡，不进行参数替换")
        elif replaceparm.find('${deposit_mb}') != -1:
            # 【会员储值占用】 ${deposit_mb}: 可用的会员卡号，
            get_res = DoMysql('CRM').get_deposit_occupy_member(type=4, groupid=str(Do_Info.group_id),
                                                               ouid=str(Do_Info.ou_id))
            res_one = replaceparm.replace('${deposit_mb}', str(get_res[0]))
            if res_one.find('${skucode}') != -1:
                # 【检验商品是否可以积分兑换】，需要替换sku_code\taticcode
                res_data = self.replace_usecheckbool(res_one)
            else:
                res_data = res_one
        elif replaceparm.find('${rel_deposti_delmb}') != -1:
            # 【根据会员占用储值GUId释放、会员储值占用】:传入已删除的会员
            get_res = DoMysql('CRM').get_deposit_occupy_member(type=2, groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id))
            res_data = replaceparm.replace('${rel_deposti_delmb}', str(get_res[0]))
        elif replaceparm.find('${rel_deposti_different_group}') != -1:
            # 【根据会员占用储值GUId释放、会员储值占用】 传入非当前集团下的会员
            get_res = DoMysql('CRM').get_deposit_occupy_member(type=3, groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id))
            res_data = replaceparm.replace('${rel_deposti_different_group}', str(get_res[0]))
        elif replaceparm.find('${tatic_code}') != -1:
            # ${tatic_code}:商品兑换策略，用于【根据兑换种类为积分兑换策略编码获取所有可以兑换的商品】
            tatic_code = DoMysql('CRM').get_integral_exchtactic(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                                type=0)
            res_data = replaceparm.replace('${tatic_code}', str(tatic_code[0]))
        elif replaceparm.find('${differ_group_tatic}') != -1:
            # ${tatic_code}:商品兑换策略，用于【根据兑换种类为积分兑换策略编码获取所有可以兑换的商品】
            tatic_code = DoMysql('CRM').get_integral_exchtactic(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                                type=3)
            res_data = replaceparm.replace('${differ_group_tatic}', str(tatic_code[0]))
        elif replaceparm.find('${relintegral_mb}') != -1:
            # 【根据会员和GUID码查询积分、卡券未释放的占用记录】 ${relintegral_mb}：积分占用的会员id
            integral_info = DoMysql('CRM').get_integral_occupy(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                               type=1)
            if integral_info != 0:
                update_one = replaceparm.replace('${relintegral_mb}', str(integral_info[1]))
                if update_one.find('${integral_respon}') != -1:
                    # 已占用的积分的guid
                    res_data = update_one.replace('${integral_respon}', str(integral_info[2]))
                else:
                    res_data = update_one
            else:
                mylog.error("查不到积分占用的会员，不进行参数替换")
        elif replaceparm.find('${coupon_mb}') != -1:
            # 卡券核销:检验卡券是否可以核销，替换会员卡号和卡券编码
            res_data = self.replace_coupon_info(replaceparm)
        elif replaceparm.find('${occpy_coupon_mb}') != -1:
            # 【积分占用及卡券占用(占用积分)】 将判断逻辑抽取到一个方法里面
            res_data = self.replace_occupy_coupon_info(replaceparm)
        elif replaceparm.find('${card_code}') != -1:
            # 验证会员密码： 正常可用的会员卡号
            card_info = DoMysql('CRM').get_member_cardcode(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id), type=0)
            res_data = replaceparm.replace('${card_code}', str(card_info[0]))
        elif replaceparm.find('${differ_group_card_code}') != -1:
            # 验证会员密码： 不在一个集团下的会员卡号
            card_info = DoMysql('CRM').get_member_cardcode(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id), type=3)
            res_data = replaceparm.replace('${differ_group_card_code}', str(card_info[0]))
        elif replaceparm.find('${del_card_code}') != -1:
            # 验证会员密码： 同一个集团，被删除的会员卡号
            card_info = DoMysql('CRM').get_member_cardcode(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id), type=1)
            res_data = replaceparm.replace('${del_card_code}', str(card_info[0]))
        elif replaceparm.find('${search_coupon_mb}') != -1:
            # 查询卡券接口
            res_data = self.replace_search_coupon_info(replaceparm)
        elif replaceparm.find('${mobile}') != -1:
            # 【发送短信获取验证码】
            mobile = DoMysql('CRM').get_member_cardcode(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id), type=2)
            res_data = replaceparm.replace('${mobile}', str(mobile[1]))
        elif replaceparm.find('${release_coupon_mb}') != -1:
            # 【撤销已占用卡券的信息】
            res_data = self.replace_release_coupon_info(replaceparm)
        elif replaceparm.find('${release_deposti_mb}') != -1:
            # 查询已占用会员储值的会员卡号和GUID，用于【根据会员占用储值GUId释放】接口
            # ${release_deposti_mb}:占用会员储值的会员卡号
            # ${rel_deposit_reponse}: 占用会员储值的GUID
            get_res = DoMysql('CRM').get_deposit_occupy_member(type=0, groupid=str(Do_Info.group_id),
                                                               ouid=str(Do_Info.ou_id))
            if get_res != 0:
                update_one = replaceparm.replace('${release_deposti_mb}', str(get_res[0]))
                if update_one.find('${rel_deposit_reponse}') != -1:
                    res_data = update_one.replace('${rel_deposit_reponse}', str(get_res[1]))
                else:
                    res_data = update_one
            else:
                mylog.error("查询到的占用会员储值的数据为空，不进行替换")
                res_data = replaceparm
        elif replaceparm.find('${had_rel_deposti_mb}') != -1:
            # 【验证占用码集合是否可用_已被释放】
            # ${had_rel_deposti_mb}: 已经释放会员储值的会员卡号
            # ${had_rel_deposti_response}：已经释放会员储值的GUID,
            get_res = DoMysql('CRM').get_deposit_occupy_member(type=1, groupid=str(Do_Info.group_id),
                                                               ouid=str(Do_Info.ou_id))
            if get_res != 0:
                update_one = replaceparm.replace('${had_rel_deposti_mb}', str(get_res[0]))
                if update_one.find('${had_rel_deposti_response}') != -1:
                    res_data = update_one.replace('${had_rel_deposti_response}', str(get_res[1]))
                else:
                    res_data = update_one
            else:
                mylog.error("查询不到数据，不进行参数替换")
                res_data = replaceparm
        elif replaceparm.find('${exist_billno}') != -1:
            # 提交消费单、上传销售记录：已经存在的消费单单号
            res_data = self.replace_exist_bill_info(replaceparm)
        else:
            res_data = replaceparm
        return res_data

    def replace_url(self, url_data):
        '''
        为了适配在不同的环境下跑，接口Url中的grouid、ouid、shopid，实现为可替换
        目前有两种的替换方式，一种直接查询数据库，将grouid、ouid、shopid保存使用；一种直接在do_info传入对应的grouid、ouid、shopid
        两种方式的区别，do_info：可以手动配置目前测试数据较多的店铺
        :param url_data:
        :return:
        '''
        # 根据配置文件的config>environment.config，配置的url中的host地址进行替换
        # 还要替换里面的集团编码、组织编码、店铺编码
        # 根据配置文件env，判断取的是那个环境，替换host的ip地址
        if Do_Info.env == "test":
            config_url = eval(ConfigRead().read_config(project_path.api_host, 'TESTHOST', 'config'))
        else:
            config_url = eval(ConfigRead().read_config(project_path.demo_api_host, 'TESTHOST', 'config'))
        # bas_info = DoMysql('BAS').get_bas_group_ou_shop_info(type=3)
        # 20211231:不采用查询数据库获取grouid ouid shopid的方式，直接从do_info中取
        data_item = string.Template(url_data)
        res_url = data_item.safe_substitute(host=str(config_url['host']), groupid=str(Do_Info.group_id),
                                            ouid=str(Do_Info.ou_id), shopid=str(Do_Info.shop_id))
        return res_url

    def replace_url_bak(self, url_data):
        '''
        为了适配在不同的化境下跑，接口Url中的grouid、ouid、shopid，实现为可替换
        目前有两种的替换方式，一种直接查询数据库，将grouid、ouid、shopid保存使用；一种直接在do_info传入对应的grouid、ouid、shopid
        两种方式的区别，do_info：可以手动配置目前测试数据较多的店铺
        :param url_data:
        :return:
        '''
        # 根据配置文件的config>environment.config，配置的url中的host地址进行替换
        # 还要替换里面的集团编码、组织编码、店铺编码
        # 根据配置文件env，判断取的是那个环境，替换host的ip地址
        if Do_Info.env == "test":
            config_url = eval(ConfigRead().read_config(project_path.api_host, 'TESTHOST', 'config'))
        else:
            config_url = eval(ConfigRead().read_config(project_path.demo_api_host, 'TESTHOST', 'config'))
        bas_info = DoMysql('BAS').get_bas_group_ou_shop_info(type=3)
        # 反射到do_info保存，可在parm查询的时候使用
        setattr(Do_Info, 'group_id', bas_info[0])
        setattr(Do_Info, 'ou_id', bas_info[2])
        # todo: 20210615：先注释掉，店铺1找不到微客数据
        # setattr(Do_Info, 'shop_id', bas_info[4])
        # print(config_url['host'])
        if url_data.find('${host}') != -1:
            update_one = url_data.replace('${host}', str(config_url['host']))
            if update_one.find('${groupid}') != -1:
                # ${groupid}: 集团编码，找到集团编码，进行替换
                GROUPID = bas_info[0]
                update_two = update_one.replace('${groupid}', str(GROUPID))
                if update_two.find('${ouid}') != -1:
                    # 替换ouid:组织id
                    OUID = bas_info[2]
                    update_three = update_two.replace('${ouid}', str(OUID))
                    if update_three.find('${shopid}') != -1:
                        # SHOPID = bas_info[4] # todo:20210615微客提交，先注释，取info里面的
                        res_url = update_three.replace('${shopid}', str(Do_Info.shop_id))
                    else:
                        res_url = update_three
                else:
                    res_url = update_two
            else:
                res_url = update_one
        else:
            res_url = url_data
        return res_url

    def replace_parm_basinfo(self, module, replaceparm):
        '''
        替换方法，针对ifpos，传入的参数要填入集团、组织、店铺
        先进行替换了集团、组织、店铺，再传相应的集团、组织和店铺，在进行过滤查询
        :param module:
        :param replaceparm:
        :return:
        '''
        if module == 'ifpos':
            # 这里不能使用string模板替换字符串的方式
            if replaceparm.find('${groupcode}') != -1 or replaceparm.find('${oucode}') != -1 \
                    or replaceparm.find('${shopcode}') != -1:
                # print(Do_Info.group_id)
                bas_info = DoMysql('BAS').get_bas_group_ou_shop_info(type=0, groupid=str(Do_Info.group_id))
                update_one = replaceparm.replace('${groupcode}', str(bas_info[1]))
                if update_one.find('${oucode}') != -1:
                    update_two = update_one.replace('${oucode}', str(bas_info[3]))
                    if update_two.find('${shopcode}') != -1:
                        res_data = update_two.replace('${shopcode}', str(bas_info[5]))
                    elif update_two.find('${differert_group_shop_code}') != -1:
                        # 传入的店铺和当前集团和组合不同
                        diff_info = DoMysql('BAS').get_bas_group_ou_shop_info(type=1, groupid=str(Do_Info.group_id))
                        res_data = update_two.replace('${differert_group_shop_code}', str(diff_info[5]))
                    else:
                        res_data = update_two
                else:
                    res_data = update_one
            else:
                res_data = replaceparm
        return res_data

    def replace_usecheckbool(self, replaceparm):
        '''
        检验商品是否可以积分兑换的skucode\taticcode
        替换逻辑：此部分用例都由条形码，替换了条形码之后，再根据兑换策略，走不同分支，进行替换
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "skucode" in param_t.keys() or "tatic_code" in param_t.keys() or "differ_group_tatic" in param_t.keys() \
                or "no_support_tatic" in param_t.keys() or "del_tatic" in param_t.keys():
                if replaceparm.find("${tatic_code}") != -1:
                    # 同一个集团下的正常的可用的策略
                    search_type = 0
                    search_txt = "同一个集团下的兑换策略"
                elif replaceparm.find("${differ_group_tatic}") != -1:
                    # 不同集团下的兑换策略，正常可用的兑换策略
                    search_type = 3
                    search_txt = "不同集团下的兑换策略"
                elif replaceparm.find("${no_support_tatic}") != -1:
                    # 不支持商品兑换的策略
                    search_type = 1
                    search_txt = "不支持商品兑换的策略"
                else:
                    # 删除
                    search_type = 2
                    search_txt = "已删除的策略"
                tatic_code = DoMysql('CRM').get_integral_exchtactic(groupid=str(Do_Info.group_id),
                                                                    ouid=str(Do_Info.ou_id), type=search_type)
                # '${skucode}:SKU编码，查看商品是否支持积分兑换用到
                sku_code = DoMysql('BAS').get_sku_info(groupid=str(Do_Info.group_id))
                if tatic_code != 0 and sku_code !=0:
                    param_t["skucode"] = str(sku_code[0])
                    param_t["tatic_code"] = str(tatic_code[0])
                    param_t["differ_group_tatic"] = str(tatic_code[0])
                    param_t["no_support_tatic"] = str(tatic_code[0])
                    param_t["del_tatic"] = str(tatic_code[0])
                    # 安全替换，只替换存在的，其他的不需要替换
                    res_data = commonFun.replace_by_stringTemple(param=replaceparm, temple=param_t)
                else:
                    mylog.exception("查询{0}_无记录，不进行参数替换".format(search_txt))
                    res_data = replaceparm
                return res_data

    def replace_coupon_info(self, replaceparm):
        '''
        卡券核销: 检验卡券是否可以核销，替换会员卡号和卡券编码
        替换逻辑：先替换了crm会员卡号，再根据场景，传入正常的、已删除、已过期的卡券进行替换
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "coupon_mb" in param_t.keys() or "couponcode" in param_t.keys() or "del_couponcode" in param_t.keys() \
                or "outdate_couponcode" in param_t.keys():
            if replaceparm.find("${couponcode}")!=-1:
                search_type = 0
                search_txt = "正常可用的卡券"
            elif replaceparm.find('${del_couponcode}') != -1:
                search_type = 1
                search_txt = "已经删除的卡券"
            else:
                search_type = 2
                search_txt = "已经过期的卡券"

        coupon_info = DoMysql('CRM').get_member_coupon(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                       shopid=str(Do_Info.shop_id), type=search_type)
        if coupon_info != 0:
            param_t["coupon_mb"] = str(coupon_info[0])
            param_t["couponcode"] = str(coupon_info[1])
            param_t["del_couponcode"] = str(coupon_info[1])
            param_t["outdate_couponcode"] = str(coupon_info[1])
            res_data = commonFun.replace_by_stringTemple(replaceparm,param_t)
        else:
            mylog.exception("查询{0}_无记录，不进行替换".format(search_txt))
            res_data =replaceparm
        return res_data
        # 旧逻辑
        # if coupon_info != 0:
        #     update_one = replaceparm.replace('${coupon_mb}', str(coupon_info[0]))
        #     if update_one.find('${couponcode}') != -1:
        #         # ${couponcode}: 正常可用的卡券
        #         res_data = update_one.replace('${couponcode}', str(coupon_info[1]))
        #     elif update_one.find('${del_couponcode}') != -1:
        #         # ${del_couponcode}： 已经删除的卡券
        #         coupon_info = DoMysql('CRM').get_member_coupon(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
        #                                                        shopid=str(Do_Info.shop_id), type=1)
        #         if coupon_info != 0:
        #             res_data = update_one.replace('${del_couponcode}', str(coupon_info[1]))
        #         else:
        #             mylog.error("卡券占用接口-查找没有占用卡券无数据，不进行参数替换！")
        #             res_data = update_one
        #     elif update_one.find('${outdate_couponcode}') != -1:
        #         # ${outdate_couponcode}：已经过期的卡券
        #         coupon_info = DoMysql('CRM').get_member_coupon(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
        #                                                        shopid=str(Do_Info.shop_id), type=2)
        #         res_data = update_one.replace('${outdate_couponcode}', str(coupon_info[1]))
        #     else:
        #         res_data = update_one
        # else:
        #     mylog.error("检验卡券是否可以核销-查找没有无数据，不进行参数替换！")
        #     res_data = replaceparm
        # return res_data

    def replace_occupy_coupon_info(self, replaceparm):
        '''
        卡券占用接口的替换参数逻辑：每个用例都含有${occpy_coupon_mb}，取这个参数作为入口
        分别替换卡券已占用的接口和卡券没有占用、卡券集合（一张已占用+一张未占用）的场景
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "coupon_code" in param_t.keys() or "occpy_coupon_mb" in param_t.keys() or "occupy_coupon" in param_t.keys() \
                or "no_coupon_code" in param_t.keys() or "had_occupy_coupon" in param_t.keys():
            if replaceparm.find("${occupy_coupon}") != -1:
                search_type = 3
                search_txt = "走卡券已被占用"
            else:
                # 走没有占用的卡券分支
                search_type = 0
                search_txt = "没有占用的卡券"


            coupon_info = DoMysql('CRM').get_member_coupon(groupid=(str(Do_Info.group_id)), ouid=str(Do_Info.ou_id),
                                                           shopid=str(Do_Info.shop_id), type=search_type)

            if coupon_info != 0:
                param_t["coupon_code"] = str(coupon_info[1])
                param_t["occupy_coupon"] = str(coupon_info[1])
                param_t["no_coupon_code"] = str(coupon_info[1])
                param_t["occpy_coupon_mb"] = str(coupon_info[0])
                DataManage().add_data_obj({"crm_memberid":str(coupon_info[3])})
                if replaceparm.find('${had_occupy_coupon}') != -1:
                    occupy_info = DoMysql('CRM').get_couponcode_by_cardcode(str(coupon_info[0]), type=1)
                    if occupy_info != 0:
                        param_t["had_occupy_coupon"] = str(occupy_info[1])
                    else:
                        mylog.exception("查找会员{0}占用卡券无记录".format(str(occupy_info[1])))

                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.exception("查找{0},无记录，不进行参数替换".format(search_txt))
                res_data = replaceparm
        return res_data
        # # 旧的替换逻辑
        # if replaceparm.find('${coupon_code}') != -1:
        #     # 走没有占用的卡券分支
        #     coupon_info = DoMysql('CRM').get_member_coupon(groupid=(str(Do_Info.group_id)), ouid=str(Do_Info.ou_id),
        #                                                    shopid=str(Do_Info.shop_id), type=0)
        #     if coupon_info != 0:
        #         # 替换没有占用的卡券分支，调用接口走占用此卡券的分支
        #         update_one = replaceparm.replace('${coupon_code}', str(coupon_info[1]))
        #         if update_one.find('${occpy_coupon_mb}') != -1:
        #             # 替换crm会员的会员卡
        #             res_data = update_one.replace('${occpy_coupon_mb}', str(coupon_info[0]))
        #         else:
        #             res_data = update_one
        #     else:
        #         mylog.error("卡券占用接口-查找没有占用卡券无数据，不进行参数替换！")
        #         res_data = replaceparm
        # elif replaceparm.find('${occupy_coupon}') != -1:
        #     # 走卡券已被占用的分支
        #     coupon_info = DoMysql('CRM').get_member_coupon(groupid=(str(Do_Info.group_id)), ouid=str(Do_Info.ou_id),
        #                                                    shopid=str(Do_Info.shop_id), type=3)
        #     if coupon_info != 0:
        #         update_one = replaceparm.replace('${occupy_coupon}', str(coupon_info[1]))
        #         if update_one.find('${occpy_coupon_mb}') != -1:
        #             res_data = update_one.replace('${occpy_coupon_mb}', str(coupon_info[0]))
        #         else:
        #             res_data = update_one
        #     else:
        #         mylog.info("卡券占用接口-查询不到卡券占用的数据，不进行参数替换")
        #         res_data = replaceparm
        # elif replaceparm.find('${no_coupon_code}') != -1 and replaceparm.find('${had_occupy_coupon}') != -1:
        #     # 走一张卡券被占用，一张卡券未被占用的分支
        #     # 先替换未被占用的
        #     coupon_info = DoMysql('CRM').get_member_coupon(groupid=(str(Do_Info.group_id)), ouid=str(Do_Info.ou_id),
        #                                                    shopid=str(Do_Info.shop_id), type=0)
        #     if coupon_info != 0:
        #         update_one = replaceparm.replace('${no_coupon_code}', str(coupon_info[1]))
        #         # 将会员卡号保存下来,用于查询同一个用户已占用的卡券
        #         setattr(Do_Info, 'card_code', coupon_info[0])
        #         if update_one.find('${occpy_coupon_mb}') != -1:
        #             update_two = update_one.replace('${occpy_coupon_mb}', str(coupon_info[0]))
        #             if update_two.find('${had_occupy_coupon}') != -1:
        #                 # 加上根据会员卡号查询用户占用的卡券，查询已占用的卡券
        #                 occupy_info = DoMysql('CRM').get_couponcode_by_cardcode(str(Do_Info.card_code), type=1)
        #                 if occupy_info != 0:
        #                     res_data = update_two.replace('${had_occupy_coupon}', str(occupy_info[1]))
        #                 else:
        #                     mylog.error("卡券占用接口-根据会员卡号，查询已占用的卡券无数据，不替换参数")
        #                     res_data = update_two
        #             else:
        #                 res_data = update_two
        #         else:
        #             res_data = update_one
        #     else:
        #         mylog.error("卡券占用接口-一张卡券被占用，一张卡券未被占用的分支，查找无数据，不做参数替换!")
        #         res_data = replaceparm
        # return res_data

    def replace_search_coupon_info(self, replaceparm):
        '''
        查看卡券的卡券信息,替换逻辑：每个用例都有${search_coupon_mb}，以这个参数作为入口，进入替换方法
        :param replaceparm:
        :return:
        '''
        # 替换查看卡券的卡券信息
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "couponcode" in param_t.keys() or "search_coupon_mb" in param_t.keys() or "tovoid_couponcode" in param_t.keys() \
                or "had_couponcode" in param_t.keys() or "outdate_couponcode" in param_t.keys() \
                or "diff_mb_couponcode" in param_t.keys():
            if replaceparm.find('${tovoid_couponcode}') != -1:
                search_type = 3
                search_txt = "已作废的卡券"
            elif replaceparm.find('${had_couponcode}') != -1:
                search_type = 1
                search_txt = "已使用的卡券"
            elif replaceparm.find('${outdate_couponcode}') != -1:
                search_type = 2
                search_txt = "已过期的卡券"
            else:
                search_type = 0
                search_txt = "没有使用过的卡券"

            coupon_info = DoMysql('CRM').search_coupon_info(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                            shopid=str(Do_Info.shop_id), type=search_type)

            if coupon_info !=0:
                param_t["couponcode"] = str(coupon_info[1])
                param_t["search_coupon_mb"] = str(coupon_info[0])
                param_t["tovoid_couponcode"] = str(coupon_info[1])
                param_t["had_couponcode"] = str(coupon_info[1])
                param_t["outdate_couponcode"] = str(coupon_info[1])

                if replaceparm.find("${diff_mb_couponcode}") !=-1:
                    # 传入当前会员，查询出非当前会员的会员卡
                    diff_mb = DoMysql('CRM').search_coupon_info(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                                shopid=str(Do_Info.shop_id),
                                                                cardcode=str(coupon_info[0]))
                    param_t["diff_mb_couponcode"] = str(diff_mb[1])

                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.exception("查询{0}，无记录，不进行参数替换".format(search_txt))
                res_data = replaceparm
        return res_data

    def replace_release_coupon_info(self, replaceparm):
        '''
        替换撤销已占用卡券的信息，根据卡券guid是否占用走不同的分支
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "coupon_respond" in param_t.keys() or "release_coupon_mb" in param_t.keys() or "no_occupy_card_respond" in param_t.keys():
            if replaceparm.find('${coupon_respond}') != -1:
                search_type = 3
                search_txt = "已被占用的卡券guid"
            else:
                search_type = 0
                search_txt = "未被占用的卡券"

            coupon_info = DoMysql('CRM').get_member_coupon(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                           shopid=str(Do_Info.shop_id), type=search_type)

            if coupon_info != 0:
                param_t["coupon_respond"] = str(coupon_info[2])
                param_t["release_coupon_mb"] = str(coupon_info[0])
                param_t["no_occupy_card_respond"] = str(coupon_info[2])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.exception("查找{0}_无记录，不进行参数替换".format(search_txt))
                res_data = replaceparm
        return res_data

    def replace_mallmeber_crmid(self, replaceparm):
        '''
        crm接口，替换商场会员id，和crm会员id,较多接口只需要传入crm会员id或商城会员id，封装到一个方法，进行替换
        :param replaceparm:
        :return:
        '''
        res = DoMysql('CRM').get_memberid(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                          shopid=str(Do_Info.shop_id), type=0)
        if replaceparm.find('${mall_memberid}') != -1 or replaceparm.find('${crm_memberid}') != -1:
            param_t = commonFun.dynamic_create_dict(replaceparm)
            if "mall_memberid" in param_t.keys() or "crm_memberid" in param_t.keys():
                param_t["mall_memberid"] = str(res[0])
                param_t["crm_memberid"] = str(res[1])
            res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
        elif replaceparm.find('${del_mall_memberid}') != -1 or replaceparm.find('${del_crm_member}') != -1:
            # 查找已经删除的会员id
            res = DoMysql('CRM').get_memberid(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                              shopid=str(Do_Info.shop_id), type=1)
            param_t = commonFun.dynamic_create_dict(replaceparm)
            if res != 0:
                if "del_mall_memberid" in param_t.keys() or "del_crm_member" in param_t.keys():
                    param_t["del_mall_memberid"] = str(res[0])
                    param_t["del_crm_member"] = str(res[1])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("查找不到已经删除的商城会员id，不进行参数替换！")
                res_data = replaceparm
        return res_data

    def replace_crm_mobile(self, replaceparm):
        '''
        修改电话接口
        :param replaceparm:
        :return:
        '''
        res = DoMysql('CRM').get_crmmemberid_and_mobile(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                        shopid=str(Do_Info.shop_id))
        mobile = CommunFun().create_mobile()
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "oldMoblie" in param_t.keys() or "crm_memberid" in param_t.keys() or "new_mobile" in param_t.keys() or "exist_mobile" in param_t.keys():
            if replaceparm.find('${exist_mobile}') != -1:
                exist_mobile = DoMysql('CRM').get_crmmemberid_and_mobile(groupid=str(Do_Info.group_id),
                                                                         ouid=str(Do_Info.ou_id),
                                                                         shopid=str(Do_Info.shop_id))
                param_t["exist_mobile"] = str(exist_mobile[1])

            param_t["oldMoblie"] = str(res[1])
            param_t["crm_memberid"] = str(res[0])
            param_t["new_mobile"] = str(mobile)
        res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
        return res_data

    def replace_pin_act_list(self, replaceparm):
        '''
        拼券列表接口
        :param replaceparm:
        :return:
        '''
        #
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "success_pin_list" in param_t.keys() or "pin_list_union" in param_t.keys() or "fail_pin_list" in param_t.keys():
            if replaceparm.find('${fail_pin_list}') != -1:
                # 拼券失败的活动id和union
                search_type = 5
            else:
                # 拼券成功的活动id和union
                search_type = 4
            pin_list_info = DoMysql('CRM').get_pinactivityid_and_teamid(groupid=str(Do_Info.group_id),
                                                                        ouid=str(Do_Info.ou_id), type=search_type)
            if pin_list_info != 0:
                param_t["fail_pin_list"] = str(pin_list_info[0])
                param_t["success_pin_list"] = str(pin_list_info[0])
                param_t["pin_list_union"] = str(pin_list_info[2])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("查不到拼券失败的拼团活动和Union,不进行参数替换！")
                res_data = replaceparm
            return res_data

    def replace_pin_act_user_list(self, replaceparm):
        '''
        获取拼团人员列表
        替换逻辑：根据不同状态的teamid来区分不同场景，来获取拼团人员列表
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "doing_teamid" in param_t.keys() or "pin_user_list_act" in param_t.keys() or "success_teamid" in param_t.keys() or "fail_teamid" in param_t.keys():
            if replaceparm.find('${doing_teamid}') != -1:
                # 拼券失败的活动id和union
                pin_act_info = DoMysql('CRM').get_pinactivityid_and_teamid(groupid=str(Do_Info.group_id),
                                                                           ouid=str(Do_Info.ou_id), type=7)
            elif replaceparm.find('${fail_teamid}') != -1:
                # 已经拼团失败的分支
                pin_act_info = DoMysql('CRM').get_pinactivityid_and_teamid(groupid=str(Do_Info.group_id),
                                                                           ouid=str(Do_Info.ou_id), type=5)
            else:
                # 拼团成功的活动id 和 团队id，用于获取拼团成员列表
                pin_act_info = DoMysql('CRM').get_pinactivityid_and_teamid(groupid=str(Do_Info.group_id),
                                                                           ouid=str(Do_Info.ou_id), type=4)
            if pin_act_info != 0:
                param_t["doing_teamid"] = str(pin_act_info[1])
                param_t["success_teamid"] = str(pin_act_info[1])
                param_t["fail_teamid"] = str(pin_act_info[1])
                param_t["pin_user_list_act"] = str(pin_act_info[0])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("查找正在拼团或拼团成功或拼团失败的活动和团队id无数据，不进行参数替换！")
                res_data = replaceparm
            return res_data

    def replace_valivate_coupon_info(self, replaceparm):
        '''
        拼团活动领券验证
        替换逻辑：根据不同的unionid区分不同的场景，可以领取、拼团活动没有券不能领取、已经领取过的卡券、活动停止等场景
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "no_getcoupon_union" in param_t.keys() or "had_getcoupon_union" in param_t.keys() or "stop_getcoupon_union" in param_t.keys() or "can_getcoupon_union" in param_t.keys() or "validate_coupon_act" in param_t.keys() or "validate_coupon_team" in param_t.keys():
            if replaceparm.find('${no_getcoupon_union}') != -1:
                # 拼团活动没有对应卡券的，不能进行领取
                validate_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                             ouid=str(Do_Info.ou_id), type=2)
            elif replaceparm.find('${had_getcoupon_union}') != -1:
                # 已经领过卡券的，不能再次进行领取
                validate_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                             ouid=str(Do_Info.ou_id), type=1)
            elif replaceparm.find('${stop_getcoupon_union}') != -1:
                # 拼券活动已经停止，不能再次进行领取
                validate_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                             ouid=str(Do_Info.ou_id), type=3)
            else:
                # 可以领取卡券的用户
                validate_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                             ouid=str(Do_Info.ou_id), type=0)
            if validate_info != 0:
                param_t["no_getcoupon_union"] = str(validate_info[2])
                param_t["can_getcoupon_union"] = str(validate_info[2])
                param_t["had_getcoupon_union"] = str(validate_info[2])
                param_t["stop_getcoupon_union"] = str(validate_info[2])
                param_t["validate_coupon_act"] = str(validate_info[0])
                param_t["validate_coupon_team"] = str(validate_info[1])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("查找不找拼团活动领券验证信息，不进行参数替换！")
                res_data = replaceparm
            return res_data

    def open_pin_activity_info(self, replaceparm):
        '''
         拼团活动开团、获取拼团详情
         替换逻辑：根据不同的unionid，区分走可以开团、已经满员的团和参加团
        :param replaceparm:
        :return:
        '''
        mylog.info("拼团活动开团的参数开始替换")
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "create_pin_unionid" in param_t.keys() or "create_pin_act" in param_t.keys() or "pin_act_coupon_id" in param_t.keys() or "join_team_id" in param_t.keys() or "full_pin_unionid" in param_t.keys() or "new_pin_unionid" in param_t.keys():
            if replaceparm.find('${full_pin_unionid}') != -1:
                # 拼团退队已经满员的拼团活动，不能参团
                act_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                        ouid=str(Do_Info.ou_id), type=4)
            elif replaceparm.find('${new_pin_unionid}') != -1 or replaceparm.find('${create_pin_unionid}') != -1:
                # 用于参加拼团活动，加入新的unionid
                # 随机生成的唯一码，用于参团
                if replaceparm.find('${new_pin_unionid}') != -1:
                    userunionid = CommunFun().create_unionid()
                    param_t['new_pin_unionid'] = userunionid
                else:
                    # 查找可以开团的拼团活动前，先手动添加一个有效的拼团活动
                    mylog.info("crm_pin_activity表插入一个有效的拼团活动")
                    # (注意：客户正式环境跑，可能会插入一下自动化的数据)
                    DoMysql('CRM').create_pin_activity(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id))
                act_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                        ouid=str(Do_Info.ou_id), type=5)
            else:
                act_info = DoMysql('CRM').get_ValidatePinActivityCoupon(groupid=str(Do_Info.group_id),
                                                                        ouid=str(Do_Info.ou_id), type=5)
            if act_info != 0:
                param_t["create_pin_unionid"] = str(act_info[2])
                param_t["full_pin_unionid"] = str(act_info[2])
                param_t["create_pin_act"] = str(act_info[0])
                param_t["pin_act_coupon_id"] = str(act_info[3])
                param_t["join_team_id"] = str(act_info[1])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("拼团活动开团或参团_查找不到可以成团的信息，不进行参数替换")
                res_data = replaceparm
            return res_data

    def replace_change_card_code(self, replaceparm):
        '''
        # 更换会员的参数替换,替换逻辑区分已绑定和未绑定的会员卡，商城会员id、会员手机号主要取当前集团下，未删除的
        :param replaceparm: 判断是否可替换的parm
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "binding_card_code" in param_t.keys() or "binding_card_mall_meb" in param_t.keys() or "binding_mobile" in param_t.keys() or "not_binding_cardcode" in param_t.keys() or "not_binding_mobile" in param_t.keys():
            if replaceparm.find('${binding_card_code}') != -1:
                # 已经绑定的会员卡
                binding_card = DoMysql('CRM').get_not_binding_cardcode(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id),
                                                                       shopid=str(Do_Info.shop_id),
                                                                       type=1)
            else:
                # 没有占用的会员卡信息
                binding_card = DoMysql('CRM').get_not_binding_cardcode(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id),
                                                                       shopid=str(Do_Info.shop_id),
                                                                       type=0)
            crm_info = DoMysql('CRM').get_memberid(groupid=str(Do_Info.group_id),
                                                   ouid=str(Do_Info.ou_id), shopid=str(Do_Info.shop_id), type=0)
            if binding_card != 0 and crm_info != 0:
                param_t["binding_card_code"] = str(binding_card[0])
                param_t["not_binding_cardcode"] = str(binding_card[0])
                param_t["binding_card_mall_meb"] = str(crm_info[0])
                # param_t["binding_mobile"] = str(crm_info[3])
                param_t["binding_mobile"] = str(binding_card[1])
                param_t["not_binding_mobile"] = str(binding_card[1])  # 没有绑定的手机号和会员卡号要对应同一个会员，不对应的情况下，会随机取一个进行判断
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("更换会员卡_查询不到数据，不进行参数替换")
                res_data = replaceparm
            return res_data

    def replace_coupon_shopissue(self, replaceparm):
        '''
        门店领券接口所需要的couponid、unionid、shopissue、issueshopid
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "use_couponId" in param_t.keys() or "use_unionid" in param_t.keys() or "shopissueid" in param_t.keys() \
                or "shopid" in param_t.keys() or "had_shopissueid" in param_t.keys() or "nopass_shopissueid" in param_t.keys() \
                or "del_shopissueid" in param_t.keys():
            if replaceparm.find('${shopissueid}') != -1 or replaceparm.find('${nopass_shopissueid}') != -1 \
                    or replaceparm.find('${del_shopissueid}') != -1:
                userunionid = CommunFun().create_unionid()
                if replaceparm.find('${shopissueid}') != -1:
                    # 走可以领券卡券的分支
                    search_type = 0
                elif replaceparm.find('${nopass_shopissueid}') != -1:
                    # 走活动没有审核通过的分支
                    search_type = 2
                else:
                    # 走活动审核通过，但活动已经删除
                    search_type = 3
                coupon_info = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                           ouid=str(Do_Info.ou_id), type=search_type)
                param_t["use_unionid"] = str(userunionid)

            elif replaceparm.find('${had_shopissueid}') != -1:
                # 走已经领取过卡券的分支
                coupon_info = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                           ouid=str(Do_Info.ou_id), type=1)
                param_t["use_unionid"] = str(coupon_info[2])

            elif replaceparm.find('${use_couponId}') != -1:
                # 已经删除的crm会员id
                if replaceparm.find('${del_crm_memberid}') != -1:
                    del_mb = DoMysql('CRM').get_memberid(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                     shopid=str(Do_Info.shop_id), type=1)
                    if del_mb != 0:
                        param_t["del_crm_memberid"] = str(del_mb[1])
                    else:
                        mylog.error("查找已经删除的crm会员失败，不进行参数替换！")
                        # res_data = replaceparm

                # 获取优惠券详情的分支
                coupon_info = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id), type=1)
                # 这里需要特殊处理param_t，其他走正常
                # 查找没删除的会员信息且会员积分大于1的用户
                memeberid = DoMysql('CRM').get_memberid(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                        shopid=str(Do_Info.shop_id), type=4)
                if coupon_info != 0 and memeberid != 0:
                    param_t["use_couponId"] = str(coupon_info[0])
                    param_t["use_unionid"] = str(coupon_info[2])
                    param_t["crm_memberid"] = str(memeberid[1])
                    param_t["mall_memberid"] = str(memeberid[0])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
                return res_data


            if coupon_info != 0:
                param_t["shopissueid"] = str(coupon_info[1])
                param_t["had_shopissueid"] = str(coupon_info[1])
                param_t["nopass_shopissueid"] = str(coupon_info[1])
                param_t["del_shopissueid"] = str(coupon_info[1])
                param_t["shopid"] = str(Do_Info.shop_id)
                param_t["use_couponId"] = str(coupon_info[0])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("门店派券_查找不到审核通过的门店派券信息，不进行参数替换！")
                res_data = replaceparm
            return res_data

    def replace_del_coupon_id(self, replaceparm):
        '''
        【卡券领券、卡券详情、根据卡券码获取使用状态】 已经删除的卡券的分支
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "del_use_couponId" in param_t.keys() or "use_unionid" in param_t.keys() or "member_id" in param_t.keys():
            del_coupon = DoMysql('CRM').get_coupon_info(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id), type=1)
            if del_coupon != 0:
                # 已经删除的卡券id
                param_t["del_use_couponId"] = str(del_coupon["id"])
                param_t["use_unionid"] = str(del_coupon["union_id"])
                param_t["member_id"] = str(del_coupon["member_id"])
                # param_t["del_use_couponId"] = str(del_coupon[0])
                # param_t["use_unionid"] = str(del_coupon[2])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("查不到已经删除的卡券，不进行参数替换！")
                res_data = replaceparm
        else:
            res_data =replaceparm
        return res_data

    def replace_send_coupon_info(self, replaceparm):
        '''
        会员领券接口 SendCouponToMemberForMall，替换卡券编码和会员信息
        替换逻辑：couponid统一设置成{send_couponId}，通过member来区分异常的场景
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "crm_memberid" in param_t.keys() or "send_couponId" in param_t.keys() or "del_crm_memberid" in param_t.keys() or "mall_memberid" in param_t.keys() or "del_coupon_crm_memberid" in param_t.keys() or "diff_group_crm_memberid" in param_t.keys() :
            if replaceparm.find('${del_crm_memberid}') != -1:
                # 传入的卡券正常，传入的会员被删除
                mb_search_type = 1
                coup_search_type = 0
            elif replaceparm.find('${del_coupon_crm_memberid}') != -1:
                # 走已经删除的卡券的分支，crm会员状态正常
                coup_search_type = 6
                mb_search_type = 0
            elif replaceparm.find('${diff_group_crm_memberid}') != -1:
                # 传入非当前组织的会员、正常可用的卡券id
                mb_search_type = 2
                coup_search_type = 0
            else:
                coup_search_type = 0
                mb_search_type = 0

            # 根据前面分支判断，决定查哪些类型的数据
            if coup_search_type == 6:
                # 查已经删除的卡券信息
                coupon_info = DoMysql('CRM').get_coupon_info(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id), type=1)
                param_t["send_couponId"] = str(coupon_info["id"])
            else:
                coupon_info = DoMysql('CRM').get_couponid_and_shopissue_id(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id), type=coup_search_type)
                param_t["send_couponId"] = str(coupon_info[0])

            # 获取没有被删除的会员信息
            member_info = DoMysql('CRM').get_memberid(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                      shopid=str(Do_Info.shop_id), type=mb_search_type)

            if coupon_info != 0 and member_info != 0:
                param_t["crm_memberid"] = str(member_info[1])
                param_t["del_crm_memberid"] = str(member_info[1])
                param_t["mall_memberid"] = str(member_info[0])
                param_t["del_coupon_crm_memberid"] = str(member_info[1])
                param_t["diff_group_crm_memberid"] = str(member_info[1])
                # param_t["send_couponId"] = str(coupon_info[0])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.exception("会员领券_查找不到符合要求的数据，不进行替换")
                res_data = replaceparm
        return res_data

    def replace_inviation_activity(self, replaceparm):
        '''
        邀请有礼记录接口和邀请活动规则接口,用于替换接口parm中的crm会员id、邀请有礼活动id
        替换逻辑：
        1. 有crm_memberid的，根据排名方式命名或邀请活动已过期，分别查询不同排名的数据进行替换
        2. 也可单独替换邀请活动id，可用于查询邀请活动规则
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "meb_crm_memberid" in param_t.keys() or "invitation_activity" in param_t.keys() or "rank_crm_memberid" in param_t.keys() or "act_crm_memberid" in param_t.keys() or "notwork_crm_memberid" in param_t.keys():
            if replaceparm.find('${meb_crm_memberid}') != -1:
                # 排名方式是：member 获取奖励类型为member
                search_type = 0
            elif replaceparm.find('${notwork_crm_memberid}') != -1:
                # 邀请活动已经失效
                search_type = 2
            else:
                # 排名方式是rank 战绩排名 \排名方式是 activity 为获得奖励
                search_type = 1

            # 根据对应的类型查询
            member_act = DoMysql('CRM').get_invitation_activity(groupid=str(Do_Info.group_id), type=search_type)

            if member_act != 0:
                param_t["meb_crm_memberid"] = str(member_act[1])
                param_t["rank_crm_memberid"] = str(member_act[1])
                param_t["act_crm_memberid"] = str(member_act[1])
                param_t["notwork_crm_memberid"] = str(member_act[1])
                param_t["invitation_activity"] = str(member_act[0])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("邀请有礼记录_查找不到符合要求的数据，不进行替换！")
                res_data = replaceparm

        return res_data

    def replace_weike_mobile(self, replaceparm):
        '''
        【微客分销：提交微客申请、查询微客审核情况、查询微客审核情况】
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "not_weike_member_mobile" in param_t.keys() or "introducermobile" in param_t.keys() or "weike_member_mobile" in param_t.keys() or "handling_weike_mobile" in param_t.keys() or "refuse_weike_mobile" in param_t.keys():
            if replaceparm.find('${not_weike_member_mobile}') != -1:
                # 还不是微客的分支
                search_type = 0
            elif replaceparm.find('${weike_member_mobile}') != -1:
                # 已经是微客的分支
                search_type = 1
            elif replaceparm.find('${handling_weike_mobile}') != -1:
                # 正在审核的微客手机号，用于查询微客的审核情况
                search_type = 3
            elif replaceparm.find('${refuse_weike_mobile}') != -1:
                # 审核状态：拒绝成为微客的手机号，用于查询微客的申请情况
                search_type = 2
            else:
                search_type = 0

            # 微客分销 提交微客申请 ${introducermobile}:微客推荐人手机号
            introducer = DoMysql('CRM').get_weike_IntroducerMobile(groupid=str(Do_Info.group_id),
                                                                   ouid=str(Do_Info.ou_id), shopid=str(Do_Info.shop_id))
            weike_info = DoMysql('CRM').not_weike_account_membermobile(groupid=str(Do_Info.group_id),
                                                                       ouid=str(Do_Info.ou_id),
                                                                       shopid=str(Do_Info.shop_id),
                                                                       type=search_type)

            if introducer != 0 and weike_info != 0:
                param_t["not_weike_member_mobile"] = str(weike_info[0])
                param_t["weike_member_mobile"] = str(weike_info[0])
                param_t["handling_weike_mobile"] = str(weike_info[0])
                param_t["refuse_weike_mobile"] = str(weike_info[0])
                param_t["introducermobile"] = str(introducer[0])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("提交微客申请、查询微客审核情况、查询微客审核情况_查找不到符合要求的数据，不进行替换！")
                res_data = replaceparm

        return res_data

    def replace_weike_expand_info(self, replaceparm):
        '''
        微客分销：查看我的收益 收益明细 提现申请
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "weike_crmid" in param_t.keys() or "expand_weike_crmid" in param_t.keys() or "no_expand_weike_crmid" in param_t.keys() or "not_weike_crmid" in param_t.keys() or "wait_check_crmid" in param_t.keys():
            if replaceparm.find('${weike_crmid}') != -1 or replaceparm.find('${not_weike_crmid}') != -1:
                # 微客的crm会员id
                if replaceparm.find('${weike_crmid}') != -1:
                    search_type = 1
                else:
                    search_type = 0

                crmid = DoMysql('CRM').not_weike_account_membermobile(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id), shopid=str(Do_Info.shop_id), type=search_type)

                if crmid != 0:
                    param_t["weike_crmid"] = str(crmid[1])
                    param_t["not_weike_crmid"] = str(crmid[1])
                    res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
                else:
                    mylog.error("微客分销我的收益、收益明细、提现申请_查找不到符合要求的数据，不进行替换！")
                    res_data = replaceparm

            elif replaceparm.find('${expand_weike_crmid}') != -1 or replaceparm.find('${no_expand_weike_crmid}') != -1 or replaceparm.find('${wait_check_crmid}') != -1:
                if replaceparm.find('${expand_weike_crmid}') != -1:
                    # 微客分销，有收益的crm会员
                    search_type = 0
                elif replaceparm.find('${no_expand_weike_crmid}') != -1:
                    # 微客分销，没有收益的crm会员
                    search_type = 1
                else:
                    # 微客分销 -提现申请存在未审核的记录
                    search_type = 2

                crmid = DoMysql('CRM').get_weike_expand_info(groupid=str(Do_Info.group_id), type=search_type)
                if crmid != 0:
                    param_t["expand_weike_crmid"] = str(crmid[0])
                    param_t["no_expand_weike_crmid"] = str(crmid[0])
                    param_t["wait_check_crmid"] = str(crmid[0])
                    res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
                else:
                    mylog.error("微客分销我的收益、收益明细、提现申请_查找不到符合要求的数据，不进行替换！")
                    res_data = replaceparm

            else:
                res_data = replaceparm

        return res_data

    def replace_bill_info(self, replaceparm):
        '''
        提交消费单的参数 包含单号、会员卡、skucode编码
        替换逻辑：正常状态的会员卡号和已经删除的会员卡号，会分别走不同分支，继续往下替换
        :param replaceparm:
        :return:
        '''
        # 随机生成的消费单单号
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "billno" in param_t.keys() or "sbillno" in param_t.keys() or "bill_skucode" in param_t.keys() \
                or "bill_card_code" in param_t.keys() or "shop_code" in param_t.keys() \
                or 'bill_del_meb' in param_t.keys() or 'no_billno_card_code' in param_t.keys():
            if replaceparm.find('${bill_card_code}') != -1 or replaceparm.find('${no_billno_card_code}'):
                # 正常状态的会员卡号
                mb_search_type = 1
            else:
                mb_search_type = 2

            # 会员卡号
            card_code = DoMysql('CRM').get_not_binding_cardcode(groupid=str(Do_Info.group_id),
                                                                ouid=str(Do_Info.ou_id),
                                                                shopid=str(Do_Info.shop_id), type=mb_search_type)

            billno = CommunFun().create_bill_no()
            sku_code = DoMysql('BAS').get_sku_info(groupid=str(Do_Info.group_id))
            shop_code = DoMysql('BAS').get_bas_group_ou_shop_info(groupid=str(Do_Info.group_id),
                                                                  ouid=str(Do_Info.ou_id),
                                                                  shopid=str(Do_Info.shop_id), type=2)

            if card_code != 0 and sku_code != 0 and shop_code != 0:
                param_t["billno"] = billno
                param_t["sbillno"] = billno
                param_t["bill_skucode"] = str(sku_code[1])
                param_t["bill_card_code"] = str(card_code[0])
                param_t["no_billno_card_code"] = str(card_code[0])
                param_t["bill_del_meb"] = str(card_code[0])
                param_t["shop_code"] = str(shop_code[5])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("提交消费单_查不到会员卡或skucode的数据，不进行编码替换")
                res_data = replaceparm

        return res_data

    def replace_exist_bill_info(self, replaceparm):
        '''
        提交消费单、上传销售记录：已经存在的消费单单号
        :param replaceparm:
        :return:
        '''
        # 替换已经存在的消费单的分支
        billno = DoMysql('CRM').get_expand_billno(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                  shopid=str(Do_Info.shop_id))
        update_one = replaceparm.replace('${exist_billno}', str(billno[0]))
        if update_one.find('${bill_card_code}') != -1:
            card_code = DoMysql('CRM').get_not_binding_cardcode(groupid=str(Do_Info.group_id),
                                                                ouid=str(Do_Info.ou_id),
                                                                shopid=str(Do_Info.shop_id), type=1)
            if card_code != 0:
                update_two = update_one.replace('${bill_card_code}', str(card_code[0]))
                if update_two.find('${bill_skucode}') != -1:
                    sku_code = DoMysql('CRM').get_sku_info(groupid=str(Do_Info.group_id))
                    if sku_code != 0:
                        # 提交消费订单，填入的是条码barcode
                        update_three = update_two.replace('${bill_skucode}', str(sku_code[1]))
                        if update_three.find('${shop_code}') != -1:
                            # 替换成云店的编码
                            shop_code = DoMysql('CRM').get_bas_group_ou_shop_info(groupid=str(Do_Info.group_id),
                                                                                  ouid=str(Do_Info.ou_id),
                                                                                  shopid=str(Do_Info.shop_id), type=2)
                            if shop_code != 0:
                                res_data = update_three.replace('${shop_code}', str(shop_code[5]))
                            else:
                                res_data = update_three
                        else:
                            res_data = update_three
                    else:
                        mylog.error("查找店铺条码失败，不进行条码替换")
                else:
                    res_data = update_two
            else:
                mylog.error("查找不到已占用的会员卡，不进行会员卡参数替换！")
                res_data = update_one
        else:
            res_data = update_one
        return res_data

    def replace_update_bill_info(self, replaceparm):
        '''
        # 修改消费单的单号、会员卡号、skucode条形码的参数替换
        :param replaceparm:
        :return:
        '''
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "update_bill_status" in param_t.keys() or "update_bill_card_code" in param_t.keys() or "update_bill_skucode" in param_t.keys():

            bill_info = DoMysql('CRM').get_sale_bill_no(groupid=str(Do_Info.group_id), ouid=str(Do_Info.ou_id),
                                                        shopid=str(Do_Info.shop_id))
            skucode = DoMysql('BAS').get_sku_by_skuid(str(bill_info[2]))
            if bill_info != 0 and skucode != 0:
                param_t["update_bill_status"] = str(bill_info[0])
                param_t["update_bill_card_code"] = str(bill_info[1])
                param_t["update_bill_skucode"] = str(skucode["barcode"])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("修改消费单_查找不到符合要求的消费单和skucode的数据，不进行参数替换")
                res_data = replaceparm
        return res_data

    def replace_check_weike_bymember(self, replaceparm):
        # 验证会员所属计佣微客
        param_t = commonFun.dynamic_create_dict(replaceparm)
        if "share_weike_id" in param_t.keys() or "new_weike_crmid" in param_t.keys() or "del_weike_crmid" in param_t.keys():
            if replaceparm.find('${new_weike_crmid}') != -1:
                # 可用的crm会员id
                search_type = 2
            else:
                # 已经删除的crmid
                search_type = 1
            crm_id = DoMysql('CRM').get_member_cardcode(groupid=str(Do_Info.group_id),
                                                        ouid=str(Do_Info.ou_id), type=search_type)
            # 可用的分享人微客id
            weike_id = DoMysql('CRM').get_weike_info(groupid=str(Do_Info.group_id),
                                                     ouid=str(Do_Info.ou_id), shopid=str(Do_Info.shop_id))
            if crm_id != 0 and weike_id != 0:
                param_t["share_weike_id"] = str(weike_id[0])
                param_t["new_weike_crmid"] = str(crm_id[2])
                param_t["del_weike_crmid"] = str(crm_id[2])
                res_data = commonFun.replace_by_stringTemple(replaceparm, param_t)
            else:
                mylog.error("验证会员所属计佣微客_查找不到可用的数据，不进行替换")
                res_data = replaceparm
        return res_data

    def get_testsuite_list(self, filepath, sheetname, apiconfig):
        '''
        根据传入的文件的sheetname和rowid，返回已经配置好的list
        :param filepath: 文件路径
        :param sheetname: excel文件的sheet_name
        :param apiconfig: apiconfig配置文件的节点名字
        :param rowid: rowid
        :return:
        '''
        # 传入具体的文件 和 sheetname，读取配置在excel文件里面的集合的list
        # 返回给config文件
        # 返回到excel读取方法，读取具体的用例的数据
        param = DoExcel(filepath, apiconfig).get_data_by_pandas()
        res = param[0]["parm"]
        return res


if __name__ == '__main__':
    wholeact_config = 'CREATEWHOLE'
    # res = DoExcel(project_path.oms_data, apiconfig).get_data_by_rowid(sheetname='oms',rowid=[2,3])

    # wholeact_data = DoExcel(project_path.promotion_data, "PROMOTIONCONFIG").get_appoint_sheet_data(sheet="test_suite")
    # res = DoExcel(project_path.test_data, "APICONFIG").get_data()
    test_data = DoExcel(project_path.oms_data, "OMS").get_data_oms()
    print(test_data)


